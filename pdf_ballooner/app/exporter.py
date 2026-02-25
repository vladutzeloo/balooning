"""PDF and CSV export logic using PyMuPDF's vector drawing API."""
from __future__ import annotations

import csv
import math
from datetime import date
from pathlib import Path

import fitz  # PyMuPDF

from app.balloon import BalloonData


def _transform_coords(bx: float, by: float, rotation: int,
                       w: float, h: float) -> tuple[float, float]:
    """Map balloon coords from the viewer's rotated space to original page space.

    *w* and *h* are the **original** (unrotated) page width and height.
    """
    if rotation == 90:
        return (w - by, bx)
    elif rotation == 180:
        return (w - bx, h - by)
    elif rotation == 270:
        return (by, h - bx)
    return (bx, by)


def export_pdf(src_path: str, dst_path: str, balloons: list[BalloonData],
               page_rotations: dict[int, int] | None = None) -> None:
    """Write a new PDF to *dst_path* with balloons drawn as vector overlays."""
    if Path(src_path).resolve() == Path(dst_path).resolve():
        raise ValueError("Cannot overwrite the original PDF. Choose a different output path.")

    if page_rotations is None:
        page_rotations = {}

    doc = fitz.open(src_path)

    for page_idx in range(len(doc)):
        page = doc[page_idx]
        rotation = page_rotations.get(page_idx, 0) % 360

        # Apply user rotation to the page so PDF viewers show it rotated.
        if rotation:
            page.set_rotation((page.rotation + rotation) % 360)

        # Original (unrotated) page dimensions for coordinate transforms.
        orig_w = page.mediabox.width
        orig_h = page.mediabox.height

        page_balloons = [b for b in balloons if b.page == page_idx]
        if not page_balloons:
            continue

        shape = page.new_shape()
        # Collect text insertions so we can draw them AFTER shape.commit(),
        # ensuring numbers render on top of the filled circles.
        text_items: list[tuple[fitz.Point, str, float, tuple]] = []

        for b in page_balloons:
            # Transform balloon coordinates from rotated viewer space
            # back to original page coordinate space.
            cx, cy = _transform_coords(
                b.balloon_center.x(), b.balloon_center.y(),
                rotation, orig_w, orig_h,
            )
            tx, ty = _transform_coords(
                b.target_point.x(), b.target_point.y(),
                rotation, orig_w, orig_h,
            )
            r  = b.diameter / 2.0
            style = b.style

            # Resolve colours and leader flag by style
            if style == "red":
                leader_color  = (0.8, 0, 0)
                circle_stroke = (0, 0, 0)
                circle_fill   = (1, 0, 0)
                text_color    = (1, 1, 1)
                draw_leader   = True
            elif style == "outline":
                leader_color  = (0.8, 0, 0)
                circle_stroke = (0, 0, 0)
                circle_fill   = None
                text_color    = (0, 0, 0)
                draw_leader   = True
            elif style == "no_arrow":
                leader_color  = (0.8, 0, 0)
                circle_stroke = (0.8, 0, 0)  # red outline
                circle_fill   = (1, 1, 1)    # white fill (not a disc)
                text_color    = (0.8, 0, 0)  # red number
                draw_leader   = False
            else:  # "default"
                leader_color  = (0.8, 0, 0)
                circle_stroke = (0, 0, 0)
                circle_fill   = (1, 1, 1)
                text_color    = (0, 0, 0)
                draw_leader   = True

            target = fitz.Point(tx, ty)
            center = fitz.Point(cx, cy)

            # --- Leader line + arrowhead ---
            if draw_leader:
                dx = tx - cx
                dy = ty - cy
                dist = math.hypot(dx, dy)
                edge = fitz.Point(cx + dx / dist * r, cy + dy / dist * r) if dist > 0 \
                       else fitz.Point(cx, cy - r)
                shape.draw_line(edge, target)
                shape.finish(color=leader_color, width=1.5, stroke_opacity=1.0)
                _draw_arrowhead(shape, edge, target, size=5, color=leader_color)

            # --- Circle ---
            shape.draw_circle(center, r)
            if circle_fill is not None:
                shape.finish(color=circle_stroke, fill=circle_fill, width=1.5, fill_opacity=1.0)
            else:
                shape.finish(color=circle_stroke, width=1.5, fill_opacity=0.0)

            # --- Collect number label for later insertion ---
            font_size = b.font_size_override if b.font_size_override > 0 else max(4.0, r * 1.1)
            text  = str(b.number)
            est_w = 0.6 * font_size * len(text)
            text_items.append((
                fitz.Point(cx - est_w / 2, cy + font_size * 0.35),
                text, font_size, text_color,
            ))

        # Commit all shapes (leaders, circles) to the page first.
        shape.commit()

        # Now insert text on top so numbers are visible over the circles.
        for pt, text, font_size, color in text_items:
            page.insert_text(
                pt, text,
                fontname="hebo",
                fontsize=font_size,
                color=color,
            )

    doc.save(dst_path, garbage=4, deflate=True)
    doc.close()


def _draw_arrowhead(shape: fitz.Shape, from_pt: fitz.Point,
                    to_pt: fitz.Point, size: float = 6,
                    color: tuple = (0.8, 0, 0)) -> None:
    dx = to_pt.x - from_pt.x
    dy = to_pt.y - from_pt.y
    length = math.hypot(dx, dy)
    if length < 1:
        return
    ux, uy = dx / length, dy / length
    px, py = -uy, ux
    half = size * 0.45
    left  = fitz.Point(to_pt.x - size * ux + half * px,
                        to_pt.y - size * uy + half * py)
    right = fitz.Point(to_pt.x - size * ux - half * px,
                        to_pt.y - size * uy - half * py)
    shape.draw_polyline([to_pt, left, right, to_pt])
    shape.finish(color=color, fill=color, width=0)


def export_csv(dst_path: str, balloons: list[BalloonData]) -> None:
    """Write balloon data to a CSV file."""
    with open(dst_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Number", "Page", "X (pts)", "Y (pts)", "Description"])
        for b in sorted(balloons, key=lambda x: (x.page, x.number)):
            writer.writerow([
                b.number,
                b.page + 1,
                round(b.balloon_center.x(), 2),
                round(b.balloon_center.y(), 2),
                b.description,
            ])


def export_excel(dst_path: str, balloons: list[BalloonData],
                 drawing_name: str = "") -> None:
    """Write a formatted inspection-sheet Excel workbook to *dst_path*."""
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Sheet"

    # ---- Palette ----
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")   # dark blue
    alt_fill   = PatternFill("solid", fgColor="D6E4F0")   # light blue
    title_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    hdr_font   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    body_font  = Font(name="Calibri", size=10)
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_bdr   = Border(left=thin_side, right=thin_side,
                        top=thin_side, bottom=thin_side)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ---- Column widths ----
    col_widths = [6, 7, 55, 22, 14, 12, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Row 1: title bar (merged A1:G1) ----
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "INSPECTION SHEET"
    title_cell.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill  = hdr_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ---- Row 2: drawing name + date ----
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Drawing: {drawing_name}" if drawing_name else "Drawing:"
    ws["A2"].font  = Font(name="Calibri", bold=True, size=10)
    ws["A2"].alignment = left

    ws.merge_cells("E2:G2")
    ws["E2"].value = f"Date: {date.today().strftime('%Y-%m-%d')}"
    ws["E2"].font  = Font(name="Calibri", size=10)
    ws["E2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ---- Row 3: blank spacer ----
    ws.row_dimensions[3].height = 6

    # ---- Row 4: column headers ----
    headers = ["#", "Page", "Characteristic / Description",
               "Nominal / Spec", "Actual", "Result", "Notes"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin_bdr
    ws.row_dimensions[4].height = 20

    # ---- Data rows (starting at row 5) ----
    sorted_bs = sorted(balloons, key=lambda b: (b.page, b.number))
    for row_idx, b in enumerate(sorted_bs, 5):
        fill = alt_fill if row_idx % 2 == 1 else None
        values = [b.number, b.page + 1, b.description or "", "", "", "", ""]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = body_font
            cell.border    = thin_bdr
            cell.alignment = center if col in (1, 2, 5, 6) else left
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 16

    # ---- Freeze panes below header ----
    ws.freeze_panes = "A5"

    wb.save(dst_path)
